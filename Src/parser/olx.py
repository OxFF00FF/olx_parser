import asyncio
import json
import os
import sys
import time
from urllib.parse import urlparse

from bs4 import BeautifulSoup as BS
from openpyxl import load_workbook
from tqdm.asyncio import tqdm_asyncio
from yarl import URL
from yaspin import yaspin

from Src.app.colors import *
from Src.app.config import app_config
from Src.app.logging_config import logger
from Src.parser.constants import limit
from Src.parser.credentials import get_token
from Src.parser.request import get_data
from Src.parser.schemas import OfferID, Region, City, Category, OffersMeta, Offer
from Src.parser.utils import open_json, format_date, save_json, get_proxy_random
from Src.tables.olx import merge_city_offers, register_styles, save_offers, process_cell


class olxParser:
    """
    Parser for OLX (https://www.olx.ua)

    API:

    - https://www.olx.ua/api/v1/offers/889972662/limited-phones/
    -
    """
    __base_url = 'https://www.olx.ua'
    __api_offers_url = f"{__base_url}/api/v1/offers"

    work_dir = os.path.join(os.path.dirname(__file__))
    data_dir = os.path.join(os.path.dirname(os.path.dirname(work_dir)), 'data')

    _cols = 150
    _bar = WHITE + '{desc} ' + LIGHT_BLUE + '| {bar} |' + LIGHT_YELLOW + ' {n_fmt}/{total_fmt} ' + DARK_GRAY + ' [Прошло: {elapsed}c · Осталось: {remaining}c · {rate_fmt}]  ' + WHITE
    _txt_all_offers = f"🔄  Парсим объявления со всех страниц  "
    _txt_numbers = f"🔄  Парсим номер телефонов  "

    def __init__(self, max_workers: int = 5, Json: bool = None, Xlsx: bool = None):
        self._workers = max_workers
        self._category_url = None

        self._save_json = Json
        self._save_xls = Xlsx

        self._semaphore = asyncio.Semaphore(self._workers)

        self.out_dir = os.path.join(self.data_dir)
        os.makedirs(self.out_dir, exist_ok=True)

    @staticmethod
    def _get_headers() -> dict:
        headers = {
            'accept': 'text/html,application/xhtml+xml,application/xml;q=0.9,image/avif,image/webp,image/apng,*/*;q=0.8,application/signed-exchange;v=b3;q=0.7',
            'accept-language': 'ru',
            'content-type': 'application/json',
            'priority': 'u=1, i',
            'sec-ch-ua': '"Not A(Brand";v="8", "Chromium";v="132", "Google Chrome";v="132"',
            'sec-ch-ua-mobile': '?0',
            'sec-ch-ua-platform': '"Windows"',
            'sec-fetch-dest': 'empty',
            'sec-fetch-mode': 'cors',
            'sec-fetch-site': 'cross-site',
            'user-agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/132.0.0.0 Safari/537.36',
        }
        return headers

    @staticmethod
    def _get_cookies() -> dict:
        cookies = {
            'lang': 'ru'
        }
        return cookies

    @staticmethod
    def _get_html(html_text: str) -> BS:
        try:
            return BS(html_text, 'html.parser')
        except:
            raise

    @staticmethod
    def _find_json(html: BS) -> dict | None:
        script_text = next((item.get_text(strip=True) for item in html.find_all('script') if item.get('id') == 'olx-init-config'), None)
        if not script_text:
            logger.warning('`script_text` not found')
            return None

        match = re.search(r'window.__PRERENDERED_STATE__= (".*?");(?:\r\n|\r|\n)', script_text)
        if not match:
            logger.warning('`__PRERENDERED_STATE__` not found')
            return None

        return json.loads(json.loads(match.group(1)))

    @staticmethod
    def _format_offer(data: dict) -> Offer:
        offer = Offer()

        price_data = next((item.get('value') for item in data.get('params') if item.get('key') == 'price'), {})

        offer.id = data.get('id')
        offer.title = data.get('title')
        offer.seller_name = data.get('contact', {}).get('name')
        offer.seller_city = data.get('location', {}).get('city', {}).get('name')
        offer.description = data.get('description', '').replace('<br />', '\n').replace('<br/>', '\n').replace('<br>', '\n')

        price_label = price_data.get('label', '')
        if '$' in price_label:
            offer.price_usd = price_data.get('value')
            offer.price_uah = price_data.get('converted_value')
        elif 'грн' in price_label:
            offer.price_uah = price_data.get('value')
            offer.price_usd = price_data.get('converted_value')
        else:
            offer.price_usd = 'n/a'
            offer.price_uah = 'n/a'

        offer.price_str = price_data.get('label')
        offer.url = data.get('url')
        offer.posted_date = format_date(data.get('created_time'))

        offer.phone_number = str(data.get('contact', {}).get('phone'))

        return offer

    async def _pagination(self, category_url: str) -> int:
        """Возвраащет пагинацию для передавнной ссылки на категорию"""
        logger.info(category_url)

        response = await self._make_request(category_url)

        html = self._get_html(response)

        script_text = next((item.get_text(strip=True) for item in html.find_all('script') if item.get('id') == 'olx-init-config'), None)
        match = re.search(r'window.__PRERENDERED_STATE__= (".*?");(?:\r\n|\r|\n)', script_text, re.DOTALL)
        if match:
            data = json.loads(json.loads(match.group(1)))
            return data.get('listing', {}).get('listing', {}).get('totalPages', 0)

    async def _make_request(self, url: str, headers: dict = None, data: dict = None, payload: dict = None, json_response: bool = None, use_proxy: bool = None) -> str | dict | None:
        """
        Делает запрос и проверяет статус ответа
        """
        headers = headers or self._get_headers()
        cookies = self._get_cookies()

        attempt, retries, delay = 0, 5, 5
        for _ in range(5):
            proxy = get_proxy_random()
            status, response = await get_data(url, headers, cookies, data, payload, proxy=proxy, Json=json_response, use_proxy=use_proxy)

            if status == 200:
                attempt += 1
                logger.debug(f"✔️  [{attempt}/{retries}] Request success. Status:{LIGHT_GREEN}{status}{WHITE}")
                return response

            elif status == 404:
                attempt += 1
                logger.debug(f"⚠️  [{attempt}/{retries}] Объявление не найдено или удалено. Status: {MAGENTA}{status}{WHITE}")
                return response if json_response else None

            elif status in (400, 401):
                attempt += 1
                logger.debug(f"⚠️  [{attempt}/{retries}] Captcha. Status: {YELLOW}{status}{WHITE}")
                await asyncio.sleep(delay)

            elif status == 500:
                attempt += 1
                if attempt == 5:
                    logger.error(f"⚠️  [{attempt}/{retries}] Не удалсоь выполнить запрос. Status: {RED}{status}{WHITE} · {response.split('See')[0]}")
                if '407' in response:
                    raise RuntimeError("Неверные данные для авторизации прокси")
                await asyncio.sleep(delay)

            else:
                attempt += 1
                try:
                    offer_id = url.strip('/').split('/')[-2]
                    offer_url = f"https://www.olx.ua/{offer_id}"

                    html = self._get_html(response)
                    tab_title = html.select_one('title').get_text()
                    if 'satisfied' in tab_title:
                        logger.debug(f"⚠️  Запрос был отклонен CloudFront. {tab_title} · {offer_url}")
                    else:
                        logger.error(f"⚠️  {tab_title} · {offer_url}")
                except:
                    logger.error(f"⚠️  Unexpected status: {status} · {url}")
                await asyncio.sleep(delay)

        return {}

    async def get_categories(self) -> list[Category]:
        """
        Получает список всех категорий
        """
        payload = {
            'operationName': 'InventoryMetadata',
            'variables': {},
            'query': 'query InventoryMetadata {\n  categories {\n    id\n    name\n    parent_id\n  }\n}',
        }
        url = 'https://production-graphql.eu-sharedservices.olxcdn.com/graphql'
        response = await self._make_request(url, payload=payload, json_response=True)
        data = response.get('data').get('categories')
        return [Category(item.get('id'), item.get('name'), 0, item.get('parent_id')) for item in data]

    async def get_items_count_for_all_categories(self, region_id: int = None, city_id: int = None, region_name: str = None, city_name: str = None, sorting_by: str = 'id') -> list[Category]:
        """
        Возвращает количество объявлений в каждой категории по ID региона или ID города
        """
        headers = {'accept-language': 'ru'}

        params = {}
        if region_id:
            params['region_id'] = region_id
        if city_id:
            params['city_id'] = city_id

        url = str(URL('https://www.olx.ua/api/v1/offers/metadata/search-categories/').with_query(params))
        response = await self._make_request(url, headers, json_response=True)
        data = response.get('data', {}).get('categories', [])

        if sorting_by == 'id':
            categories = sorted(data, key=lambda item: item.get('id', 0))
        elif sorting_by == 'count':
            categories = sorted(data, key=lambda item: item.get('count', 0))
        else:
            categories = data

        if self._save_json:
            save_json(data, os.path.join(self.out_dir, f'{region_id}_{region_name}_{city_name}__categories.json'))
        return [Category(item['id'], None, item['count'], 0) for item in categories]

    async def get_category_info(self, url) -> Category:
        """
        Получает ID и TITLE категории по URL (https://www.olx.ua/nedvizhimost/kvartiry/)
        """
        breadcrumb = urlparse(url).path.strip('/').replace('/', ',')
        url = f"https://www.olx.ua/api/v1/friendly-links/query-params/{breadcrumb}"
        data = await self._make_request(url, json_response=True)

        category_id = data.get('data').get('category_id')
        title = data.get('metadata').get('seo').get('title')

        return Category(category_id, title, 0, 0)

    async def _get_offer_id(self, url: str) -> OfferID:
        """
        Получает и возвращает ID объявления из переданной ссылки
         - Ссылка объявление https://www.olx.ua/d/uk/obyavlenie/vdeokarta-pny-geforce-rtx-4060-ti-8gb-verto-vcg4060t8dfxpb1-o-IDWYtNQ.html

        :param url: Ссылка
        :return: Возвращает ID Объявления
        """
        response = await self._make_request(url)

        html = self._get_html(response)
        script_text = next((item.get_text(strip=True) for item in html.find_all('script') if '@type' in item.text), None)

        try:
            data = json.loads(script_text)
            ad_id = data.get('sku')
            return OfferID(value=ad_id)
        except Exception as e:
            logger.error(f"Failed to get ad_id. Error: {e} · {url}")

    async def get_regions(self, sorting_by: str = 'id') -> list[Region]:
        """
        Возвращает сортированный по ID список регионов

        :param sorting_by: Сортировка (id, name)
        """
        url = 'https://www.olx.ua/api/v1/geo-encoder/regions/'

        response = await self._make_request(url, json_response=True)
        data = response.get('data', [])

        if sorting_by == 'id':
            regions = sorted([(item.get('id'), item.get('name')) for item in data], key=lambda x: x[0])
        elif sorting_by == 'name':
            regions = sorted([(item.get('id'), item.get('name')) for item in data], key=lambda x: x[1])
        else:
            regions = [(item.get('id'), item.get('name')) for item in data]

        if self._save_json:
            save_json(data, os.path.join(self.out_dir, 'olx__regions.json'))
        return [Region(id=item[0], name=item[1]) for item in regions]

    async def get_cities(self, region: Region, sorting_by='id') -> list[City]:
        """
        Возвращает список ID городов выбранного регона

        :param region: ID региона 1-25
        :param sorting_by: Сортировка (id, name)
        """
        url = f'https://www.olx.ua/api/v1/geo-encoder/regions/{region.id}/cities/?limit=5000'

        response = await self._make_request(url, json_response=True)
        data = response.get('data')

        if sorting_by == 'id':
            cities = sorted([(item.get('id'), item.get('name')) for item in data], key=lambda x: x[0])
        elif sorting_by == 'name':
            cities = sorted([(item.get('id'), item.get('name')) for item in data], key=lambda x: x[1])
        else:
            cities = [(item.get('id'), item.get('name')) for item in data]

        if self._save_json:
            save_json(data, os.path.join(self.out_dir, f'{region.id}_{region.name}__cities.json'))
        return [City(id=item[0], name=item[1]) for item in cities]

    async def get_offers_count(self, category_id: int, region_id: int = None, city_id: int = None, facet_field: str = 'region') -> OffersMeta:
        """
        Получает общее, видимое и количество объявлений в регионах по ID Категории. И дополнительно по ID Региона и ID Города

        :param city_id: ID Города
        :param region_id: ID Региона
        :param category_id: ID Категории
        :param facet_field: Где показывать количество, по регионам или районам (region,  district)
        """
        headers = {'accept-language': 'ru'}

        facet_data = json.loads('[{"field":"region","fetchLabel":true,"fetchUrl":true,"limit":199}]')[0]
        facet_data['field'] = facet_field

        params = {
            'offset': '0',
            'limit': '40',
            'category_id': category_id,
            'filter_refiners': 'spell_checker',
            'facets': json.dumps([facet_data]),
        }

        if region_id:
            params['region_id'] = region_id
        if city_id:
            params['city_id'] = city_id

        url = str(URL('https://www.olx.ua/api/v1/offers/metadata/search/').with_query(params))
        response = await self._make_request(url, headers, json_response=True)
        data = response.get('data', [])

        regions = [
            Region(id=item.get('id'), count=item.get('count'), name=item.get('label'), url=f"https://www.olx.ua/{item.get('url').strip('/')}")
            for item
            in data.get('facets', {}).get(facet_field, [])
        ]

        if self._save_json:
            save_json(data, os.path.join(self.out_dir, f'category_{category_id}_{region_id}_{city_id}__offers_count.json'))
        return OffersMeta(data.get('visible_total_count'), data.get('total_count'), regions)

    async def get_category_name(self, category_id: int) -> str:
        """Возвращает название и путь категории. Например `Главная > Недвижимость > Квартиры` по переданному ID категории"""
        params = {
            'params[category_id]': category_id,
            'page': 'ads',
            'params[offset]': '0',
            'params[limit]': '40',
            'params[currency]': 'UAH',
            'params[filter_refiners]': 'spell_checker',
            'params[viewType]': 'gallery',
            'dfp_user_id': '0',
            'advertising_test_token': '',
        }

        url = str(URL('https://www.olx.ua/api/v1/targeting/data/').with_query(params))
        response = await self._make_request(url, json_response=True)
        targeting = response.get('data', []).get('targeting', {})
        return ' > '.join([v for k, v in targeting.items() if 'name' in k])

    async def get_offers_from_page(self, category_url: str) -> list[dict]:
        """
        Получает список объявлений со всех страниц категории(ссылки).

        1. Определяется количество страниц в категории.
        2. Для каждой страницы извлекает информацию об объявлениях.
        3. Парсит данные с помощью BeautifulSoup и регулярных выражений.
        4. Возвращает список словарей с данными по каждому объявлению (идентификатор, URL, заголовок и контакт).

        :param category_url: URL категории, с которой необходимо получить предложения.
                             Должен быть корректным URL для страницы с товарами или объявлениями.
        :return: Список словарей, каждый из которых содержит информацию о предложении.
                 Каждый словарь имеет следующие ключи:
                    - 'id' (str): Уникальный идентификатор объявления.
                    - 'url' (str): URL страницы объявления.
                    - 'title' (str): Заголовок объявления.
                    - 'contact' (dict): Контактная информация (если доступна) из объявления.

        Пример возвращаемого результата:
        [
            {
                'id': '868972409',
                'url': 'https://www.olx.ua/api/v1/offers/868972409',
                'title': 'Продается видеокарта',
                'contact': {
                    "name": "AMAZIN інтернет магазин",
                    "phone": true,
                    "chat": true,
                    "negotiation": true,
                    "courier": true
                }
            },
            ...
        ]
        """

        result = []
        self._category_url = category_url

        total_pages = await self._pagination(category_url)
        tasks = [
            self._make_request(f'{category_url}/?page={page + 1}')
            for page
            in range(total_pages)
        ]

        results = await tqdm_asyncio.gather(*tasks, desc=self._txt_all_offers, bar_format=self._bar, ncols=self._cols, dynamic_ncols=True, leave=False, ascii=' ▱▰')
        for response in results:
            html = self._get_html(response)
            data = self._find_json(html)

            products = data.get('listing', {}).get('listing', {}).get('ads', [])
            for n, product in enumerate(products):
                title = product.get('title')
                url = product.get('url')
                result.append(dict(
                    id=product.get('id'),
                    title=title,
                    url=url,
                    contact=product.get('contact'))
                )
                logger.debug(f"[{n}/{len(products)}] |  {title[:50].ljust(50)}  ·  {url}")

        if self._save_json:
            await save_json(result, os.path.join(self.out_dir, 'urls.json'))
        return result

    async def _offers_from_first_page(self, category_id: int, region_id: int = None, city_id: int = None) -> tuple[list, str | None]:
        """
        Получает объявления с первой страницы через API и возвращает ссылку на следующую страницу (если есть).

        Формирует запрос с параметрами фильтрации по категории, региону и городу, отправляет его,
        извлекает список объявлений и ссылку на следующую страницу из ответа.

        :param category_id: Идентификатор категории.
        :param region_id: (Необязательный) ID региона.
        :param city_id: (Необязательный) ID города.

        :return: Кортеж из списка объявлений (сырой формат) и URL следующей страницы (или None).
        """
        params = {
            'offset': '0',
            'limit': '40',
            'category_id': category_id,
            'currency': 'UAH',
            'filter_refiners': 'spell_checker',
            'facets': '[{"field":"district","fetchLabel":true,"fetchUrl":true,"limit":30}]',
        }
        if region_id:
            params['region_id'] = region_id
        if city_id:
            params['city_id'] = city_id

        url = str(URL(self.__api_offers_url).with_query(params))
        response = await self._make_request(url, json_response=True)
        offers = response.get('data')

        next_page = response.get('links', {}).get('next', {}).get('href')
        return offers, next_page

    async def get_offers_from_api(self, category_id: int, region_id: int = None, city_id: int = None) -> list[Offer]:
        """
        Получает все объявления из API, проходя по всем страницам результата.

        Выполняет первый запрос к API и получает ссылки на следующие страницы.
        Загружает остальные страницы параллельно, собирает все результаты и форматирует их в список объектов `Offer`.

        :param category_id: Идентификатор категории товаров.
        :param region_id: (Необязательный) Идентификатор региона.
        :param city_id: (Необязательный) Идентификатор города.

        :return: Список отформатированных объявлений (`Offer`).
        """
        all_offers_raw = []
        page = 0

        offers, next_page = await self._offers_from_first_page(category_id, region_id, city_id)
        all_offers_raw.extend(offers or [])

        page_urls = []
        while next_page:
            page += 1
            page_urls.append(next_page)
            response = await self._make_request(next_page, json_response=True)
            next_page = response.get('links', {}).get('next', {}).get('href')
            print(f"\r🔄  Страница {page}", end='', flush=True)

        tasks = [
            self._make_request(url, json_response=True)
            for url
            in page_urls
        ]
        results = await tqdm_asyncio.gather(*tasks, desc=self._txt_all_offers, bar_format=self._bar, ncols=self._cols, dynamic_ncols=True, leave=False, ascii=' ▱▰')

        for response in results:
            all_offers_raw.extend(response.get('data', []))

        return [self._format_offer(offer) for offer in all_offers_raw]

    async def get_offers_from_graphql(self, page: int = None, category_id: int = None, region_id: int = None, city_id: int = None, currency: str = None) -> dict:
        """
        Получает список объявлений через GraphQL API с пагинацией и фильтрацией.

        Формирует GraphQL-запрос с параметрами поиска: категория, регион, город, валюта и страница.
        Отправляет запрос и возвращает данные объявлений из ответа.
        При включённом `_save_json` сохраняет ответ в JSON-файл.

        :param page: Номер страницы (начинается с 1). По умолчанию 1.
        :param category_id: Идентификатор категории для фильтрации.
        :param region_id: Идентификатор региона.
        :param city_id: Идентификатор города.
        :param currency: Валюта ('USD' или 'UAH').

        :return: Список объявлений (поле `data` из ответа GraphQL) или пустой словарь, если данных нет.
        """
        if page is None or page < 1:
            page = 1
        offset = (page - 1) * limit

        params = [{"key": "limit", "value": str(limit)}, {"key": "offset", "value": str(offset)}]

        if category_id:
            params.append({"key": "category_id", "value": category_id})
        if category_id:
            params.append({"key": "region_id", "value": region_id})
        if category_id:
            params.append({"key": "city_id", "value": city_id})
        if currency:
            if currency.lower() == 'usd':
                params.append({"key": "currency", "value": "USD"})
            if currency.lower() == 'uah':
                params.append({"key": "currency", "value": "UAH"})

        payload = {
            "query": """
                query ListingSearchQuery($searchParameters: [SearchParameter!] = {key: "", value: ""}) {
                  clientCompatibleListings(searchParameters: $searchParameters) {
                    ... on ListingSuccess {
                      metadata { total_elements visible_total_count }
                      links { self { href } }
                      data {
                        id
                        title
                        status
                        url
                        description
                        protect_phone
                        contact { courier chat name negotiation phone }
                        category { id type }
                        location { city { id name normalized_name } district { id name } region { id name } }
                        params { key value { ... on PriceParam { value currency converted_value converted_currency label } } }
                      }
                    }
                  }
                }
            """,
            "variables": {
                "searchParameters": params
            }
        }

        url = 'https://www.olx.ua/apigateway/graphql'
        response = await self._make_request(url, payload=payload, json_response=True)
        data = response.get('data', {})

        if self._save_json:
            save_json(data, os.path.join(self.out_dir, f'{category_id}_{region_id}_{city_id}__offers_graphql.json'))
        return data.get('clientCompatibleListings', {}).get('data', {})

    async def get_phone_number(self, ad_id: OfferID, use_proxy: bool = None, response_only: bool = None) -> str | dict | None:
        """
        Асинхронно получает номера телефонов для объявления по его ID через API.

        Формирует запрос с необходимыми заголовками, включая токен авторизации.
        При ошибке пытается обновить токен и повторить запрос.
        Если `response_only` установлен, возвращает полный JSON-ответ, иначе — объединённые номера телефонов в строку.

        :param ad_id: Идентификатор объявления.
        :param use_proxy: Флаг использования прокси для запроса.
        :param response_only: Если True — возвращает полный ответ API вместо строкового номера.

        :return: Строка с номерами телефонов через " · ", или словарь с данными ответа, или None при ошибке.
        """
        headers = {
            'accept': 'text/html,application/xhtml+xml,application/xml;q=0.9,image/avif,image/webp,image/apng,*/*;q=0.8,application/signed-exchange;v=b3;q=0.7',
            'accept-language': 'ru-RU,ru;q=0.9,en-US;q=0.8,en;q=0.7',
            'authorization': get_token(),
            'cache-control': 'no-cache',
            'pragma': 'no-cache',
            'priority': 'u=0, i',
            'sec-ch-ua': '"Google Chrome";v="137", "Chromium";v="137", "Not/A)Brand";v="24"',
            'sec-ch-ua-mobile': '?0',
            'sec-ch-ua-platform': '"Windows"',
            'sec-fetch-dest': 'document',
            'sec-fetch-mode': 'navigate',
            'sec-fetch-site': 'none',
            'sec-fetch-user': '?1',
            'upgrade-insecure-requests': '1',
            'user-agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/137.0.0.0 Safari/537.36',
        }

        phones = []
        url = f'{self.__base_url}/api/v1/offers/{ad_id}/limited-phones/'

        try:
            async with self._semaphore:
                data = await self._make_request(url, headers, json_response=True, use_proxy=use_proxy)
                if response_only:
                    return data

                if 'error' in data:
                    error = data.get('error', {})
                    error_detail = error.get('detail')

                    if error == 'invalid_token' or error_detail == 'Disallowed for this user':
                        headers['authorization'] = get_token()
                        data_2 = await self._make_request(url, headers, json_response=True, use_proxy=use_proxy)
                        phones = data_2.get('data', {}).get('phones', [])

                    else:
                        logger.debug(f"⛔  Failed to get phone_numbers: {data}")

                else:
                    phones = data.get('data', {}).get('phones', [])

                return ' · '.join([str(p) for p in phones]) if phones else ''

        except:
            logger.debug(f"⚠️  Failed to get phone_numbers: {self.__api_offers_url}/{ad_id}")
            return None

    async def parse_phones_from_file(self, filename, show_info=None):
        """
        Парсит номера телефонов из указанного Excel-файла.

        Открывает файл, считывает данные с первого листа начиная со второй строки,
        параллельно обрабатывает каждую строку (через `process_cell`), обновляет книгу и сохраняет изменения.
        По завершении переименовывает файл, выводит информацию и предлагает пользователю завершить или перезапустить процесс.

        :param filename: Имя Excel-файла для обработки (в директории `data`).
        :param show_info: Флаг для вывода информационных сообщений и прогресса.
        """
        if show_info:
            logger.info('🔄  Начинается сбор номеров из файла')
            if app_config.USE_PROXY:
                logger.info(f"🟢  Использование прокси {LIGHT_GREEN}ВКЛЮЧЕНО{WHITE}")
            else:
                logger.info(f"🔴  Использование прокси {LIGHT_RED}ОТКЛЮЧЕНО{WHITE}")
            time.sleep(3)

        # Путь до merged таблицы
        wb_path = os.path.join(self.data_dir, filename)

        # Открываем файл и делаем первую страницу активной
        with yaspin(text="Чтение") as spinner:
            wb = load_workbook(wb_path)
            ws = wb.active
            if show_info:
                spinner.text = 'Готово'
                spinner.ok('✔️')

        # Добавляем стили ячеек в книгу
        register_styles(wb)

        # Получаем офферы из таблицы начиная со второй строки и до конца в виде списка
        offers_data = list(ws.iter_rows(min_row=2, values_only=True))
        counter = {'value': 0}

        tasks = [
            process_cell(self, n, item, len(offers_data), counter, ws, wb, wb_path)
            for n, item
            in enumerate(offers_data)
        ]

        for coro in asyncio.as_completed(tasks):
            await coro

        with yaspin(text="Сохранение") as spinner:
            wb.save(wb_path)
            if show_info:
                spinner.text = 'Сохранено'
                spinner.ok('✔️')

            completed_wb_path = os.path.join(os.path.dirname(wb_path), f"+ {os.path.basename(wb_path)}")
            os.rename(wb_path, completed_wb_path)

        if show_info:
            print('\n[процесс завершил работу с кодом 0]')
            while True:
                answer = input(f"Теперь вы можете закрыть этот терминал с помощью клавиши {UNDERLINED}Q{RESET}{WHITE}. Или нажмите клавишу {UNDERLINED}ENTER{RESET}{WHITE} для перезапуска.")
                if answer.lower() == 'q' or answer.lower() == 'й':
                    break
                os.startfile(os.path.join(self.data_dir, os.path.dirname(wb_path)))
                os.system("cls")
                os.execl(sys.executable, sys.executable, *sys.argv)

    async def run(self, region_id=None, city_id=None):
        """
        Запускает парсер объявлений по регионам, городам и категориям.

        Если заданы `region_id` и `city_id`, обрабатывает только указанный регион и город.
        Иначе последовательно обрабатывает все регионы и города, при этом показывает прогресс с помощью файла `last_indexes.json`.
        Для каждого региона и города собирает объявления по категориям, сохраняет их и объединяет таблицы.
        По завершении открывает папку с результатами и позволяет перезапустить или завершить программу через консоль.

        :param region_id: (необязательно) Идентификатор региона для обработки.
        :param city_id: (необязательно) Идентификатор города для обработки.
        """
        help_message = f"\n{'─' * 30}| 📰  {BOLD}{LIGHT_MAGENTA}Найдено{RESET} / 📚  {BOLD}{LIGHT_CYAN}Страниц{RESET} / 📥  {BOLD}{RED}Собрано{RESET}{WHITE} / 📦  Всего собрано |{'─' * 30}"

        os.system('cls')
        logger.info('ℹ️  Начинается сбор объявлений для выбранного региона и города')
        time.sleep(1)

        total_collected = 0
        indexes_path = os.path.join(self.data_dir, 'last_indexes.json')

        # Получаем сохранённые индексы (если есть)
        indexes = open_json(indexes_path) if os.path.exists(indexes_path) else {"region": 0, "city": 0, "category": 0}
        save_json(indexes, indexes_path)

        regions = await self.get_regions()
        if region_id is not None:
            regions = [r for r in regions if r.id == region_id]
            indexes["region"] = 0  # сбрасываем прогресс, чтобы начать с начала выбранного региона

        for n_region, region in enumerate(regions):
            logger.debug(repr(region))
            if n_region < indexes["region"]:
                continue  # Пропускаем уже обработанные регионы

            print(f"\n{LIGHT_BLUE}[{n_region + 1} / {len(regions)}]{WHITE} |  Регион:  {LIGHT_YELLOW}{region.name.ljust(20)}{WHITE}  🆔  {region.id}")

            cities = await self.get_cities(region)
            if city_id is not None:
                cities = [c for c in cities if c.id == city_id]
                indexes["city"] = 0  # сбрасываем прогресс по городам, чтобы начать с начала выбранного города

            for n_city, city in enumerate(cities):
                logger.debug(repr(city))
                if n_region == indexes["region"] and n_city < indexes["city"]:
                    continue  # Пропускаем уже обработанные города

                print(f"{LIGHT_BLUE}[{n_city + 1} / {len(cities)}]{WHITE} |  Город:   {LIGHT_YELLOW}{city.name.ljust(20)}{WHITE}  🆔  {city.id}", end="")

                categories = await self.get_items_count_for_all_categories(region.id, city.id, region.name, city.name)

                if not categories:
                    print(" | Объявлений не найдено")
                    input(f"Нажмите {UNDERLINED}ENTER{RESET}{WHITE} для перезапуска")
                    os.execl(sys.executable, sys.executable, *sys.argv)
                    exit()
                else:
                    print(help_message)
                    for n_category, category in enumerate(categories):
                        logger.debug(repr(category))
                        if n_region == indexes["region"] and n_city == indexes["city"] and n_category < indexes["category"]:
                            continue  # Пропускаем уже обработанные категории

                        category_name = await self.get_category_name(category.id)
                        offers = await self.get_offers_from_api(category.id, region.id, city.id)
                        save_offers(offers, region_id, region.name, city.id, city.name, category.id, category_name, self.out_dir, self._save_json, self._save_xls)

                        offers_count = await self.get_offers_count(category.id, region.id, city.id)
                        total_pages = (offers_count.visible_total + limit - 1) // limit
                        max_offers = offers_count.visible_total

                        total_collected += max_offers
                        print(f"{LIGHT_BLUE}[{n_category + 1} / {len(categories)}]{WHITE} |   🆔  {category.id} · {YELLOW}{category_name[:80].ljust(80)}{WHITE} | "
                              f"📰  {BOLD}{LIGHT_MAGENTA}{offers_count.total}{RESET} / "
                              f"📚  {BOLD}{LIGHT_CYAN}{total_pages}{RESET}{WHITE} / "
                              f"📥  {BOLD}{RED}{max_offers}{RESET}{WHITE} / "
                              f"📦  {total_collected}")
                        save_json({"region": n_region, "city": n_city, "category": n_category + 1}, indexes_path)

                    print(help_message.strip())
                    time.sleep(1)
                    print(f"✅  Сбор объявлений по всем категориям в {LIGHT_YELLOW}{region.name}{WHITE} города {LIGHT_YELLOW}{city.name}{WHITE} завершен")
                    merge_city_offers(self.data_dir, region.name, region.id, city.name, city.id, self._bar)

                break

            indexes["city"] = 0
            indexes["category"] = 0
            save_json({"region": n_region, "city": 0, "category": 0}, indexes_path)

            break

        print(f"✅  Парсинг завершён · Всего собрано объявлений: {BOLD}{total_collected}{RESET}{WHITE}")

        print('\n[процесс завершил работу с кодом 0]')
        while True:
            answer = input(f"Теперь вы можете закрыть этот терминал с помощью клавиши {UNDERLINED}Q{RESET}{WHITE}. Или нажмите {UNDERLINED}ENTER{RESET}{WHITE} для перезапуска")
            if answer.lower() == 'q' or answer.lower() == 'й':
                break
            os.startfile(self.data_dir)
            os.system("cls")
            os.execl(sys.executable, sys.executable, *sys.argv)
