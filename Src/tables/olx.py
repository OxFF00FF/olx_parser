import asyncio
import os
import time

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, Alignment, NamedStyle, PatternFill
from openpyxl.utils import get_column_letter
from tqdm import tqdm
from yaspin import yaspin

from Src.app.colors import *
from Src.app.logging_config import logger
from Src.parser.credentials import get_token
from Src.parser.schemas import Offer
from Src.parser.utils import validate_filename, clickable_file_link

lock = asyncio.Lock()

green_fill = PatternFill(start_color="47ff94", end_color="47ff94", fill_type="solid")
orange_fill = PatternFill(start_color="ff6347", end_color="ff6347", fill_type="solid")
lavender_fill = PatternFill(start_color="baa4fc", end_color="baa4fc", fill_type="solid")
gray_fill = PatternFill(start_color="bbbbbb", end_color="bbbbbb", fill_type="solid")

hlink_style = Font(color="0000FF", underline="single")

number_style = NamedStyle(name="number_style", number_format="0", alignment=Alignment(horizontal='left'))
success_status = NamedStyle(name="success_status", fill=green_fill, alignment=Alignment(horizontal='center'))
error_status = NamedStyle(name="error_status", fill=orange_fill, alignment=Alignment(horizontal='center'))
hidden_status = NamedStyle(name="hidden_status", fill=lavender_fill, alignment=Alignment(horizontal='center'))
not_specified_status = NamedStyle(name="not_specified_status", fill=gray_fill, alignment=Alignment(horizontal='center'))


def save_offers_excel(content: list[Offer], filepath: str, show_info: bool = True):
    """
    Сохраняет список предложений (`Offer`) в Excel-файл с заголовками, форматированием и гиперссылками.

    Если файл уже существует — данные добавляются в конец существующего листа.
    При первом создании файла добавляются заголовки с форматированием.
    Каждая строка данных включает гиперссылки на название объявления и URL.

    :param content: Список предложений для сохранения в таблице.
    :param filepath: Путь к файлу Excel, куда сохраняются данные.
    :param show_info: Флаг, указывающий, логировать ли информацию об успешном сохранении.
    """
    column_mapping = {
        'id': 'ID',
        'title': 'Объявление',
        'number': 'Номер телефона',
        'selelr_name': 'Продавец',
        'seller_city': 'Город',
        'description': 'Описание',
        'price_usd': 'Стоимость USD',
        'price_uah': 'Стоимость UAH',
        'price_text': 'Цена формат.',
        'posted_date': 'Дата публикации',
        'url': 'Ссылка',
    }

    # Добавление заголовков
    headers = [column_mapping[key] for key in column_mapping.keys()]

    if os.path.exists(filepath):
        wb = load_workbook(filepath)
        ws = wb.active
    else:
        wb = Workbook()
        ws = wb.active
        ws.title = "Объявления"
        ws.append(headers)
        for col_num, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_num)
            cell.font = Font(bold=True)
            cell.alignment = Alignment(horizontal='center')

    # Словарь для расчета ширины колонок
    column_widths = {i: len(headers[i - 1]) + 2 for i in range(1, len(headers) + 1)}

    unique_offers_by_id = []
    seen_ids = set()

    for offer in content:
        if offer.id not in seen_ids:
            unique_offers_by_id.append(offer)
            seen_ids.add(offer.id)

    for offer in unique_offers_by_id:
        logger.debug(f"📦  {repr(offer)}")

        row = [
            offer.id,
            offer.title,
            offer.phone_number,
            offer.seller_name or '',
            offer.seller_city or ' ',
            offer.description or ' ',
            offer.price_usd or ' ',
            offer.price_uah or ' ',
            offer.price_str or ' ',
            offer.posted_date or '',
            offer.url or '',
        ]
        ws.append(row)

        # Обновляем максимальную ширину
        for col_num, value in enumerate(row, 1):
            val_str = str(value)
            column_widths[col_num] = max(column_widths[col_num], len(val_str) + 10)
            cell = ws.cell(row=ws.max_row, column=col_num)
            cell.alignment = Alignment(horizontal='left')

        # создание ссылки из названия товара
        offer_title_cell = ws.cell(row=ws.max_row, column=2)
        offer_title_cell.hyperlink = offer.url
        offer_title_cell.font = hlink_style

        # создание ссылки для url
        offer_url_cell = ws.cell(row=ws.max_row, column=11)
        offer_url_cell.hyperlink = offer.url
        offer_url_cell.font = hlink_style

    # Установка ширины колонок
    for col_num, width in column_widths.items():
        ws.column_dimensions[get_column_letter(col_num)].width = min(width, 100)

    try:
        wb.save(filepath)
        if show_info:
            print(f"💾  {LIGHT_GREEN}Файл сохранен в `{clickable_file_link(filepath)}` {WHITE}")
    except PermissionError as e:
        logger.error(f"Не удалось сохранить EXCEL файл: {e}")


def merge_city_offers(data_dir: str, region_name: str, region_id: int, city_name: str, city_id: int, bar: str):
    """
    Объединяет все XLSX-файлы одного города в один Excel-файл с форматированием и гиперссылками.

    Проходит по всем Excel-файлам указанного города, считывает данные, объединяет их в одну итоговую таблицу,
    добавляет гиперссылки на объявления и применяет форматирование. Итоговый файл сохраняется в директорию `data_dir`.

    :param data_dir: Базовая директория, содержащая XLSX-файлы по регионам и городам.
    :param region_name: Название региона.
    :param region_id: Идентификатор региона.
    :param city_name: Название города.
    :param city_id: Идентификатор города.
    :param bar: Строка формата для отображения прогресс-бара (используется в tqdm).
    """
    print("🔄  Объединение таблиц")
    time.sleep(1)

    xlsx_path = os.path.join(data_dir, f"{region_name.replace(' ', '-')}_{region_id}", f"{city_name}_{city_id}")
    save_path = os.path.join(data_dir, f"merged_{region_name}_{region_id}__{city_name}_{city_id}.xlsx")

    merged_wb = Workbook()
    output_ws = merged_wb.active
    output_ws.title = f"{region_name} · {city_name}"
    header_written = False
    column_widths = {}
    seen_ids = set()

    xlsx_files = [f for f in os.listdir(xlsx_path) if f.endswith('xlsx')]

    progress_bar = tqdm(xlsx_files, bar_format=bar, ncols=150, leave=False, ascii=' ▱▰')

    for filename in progress_bar:
        file_path = os.path.join(xlsx_path, filename)
        progress_bar.set_description_str(f"{filename[:50]}...")

        processed_wb = load_workbook(file_path)
        processed_ws = processed_wb.active
        rows = list(processed_ws.iter_rows(values_only=True))

        if not rows:
            continue

        # Запись заголовка с форматированием
        if not header_written:
            output_ws.append(rows[0])
            for col_num, header in enumerate(rows[0], 1):
                cell = output_ws.cell(row=1, column=col_num)
                cell.font = Font(bold=True)
                cell.alignment = Alignment(horizontal='center')
                column_widths[col_num] = len(str(header)) + 2
            header_written = True
            data_rows = rows[1:]
        else:
            data_rows = rows[1:]

        for row in data_rows:
            offer_id = row[0]
            if offer_id in seen_ids:
                continue
            seen_ids.add(offer_id)
            output_ws.append(row)

    print("\r✅  Объединение завершено", end="", flush=True)
    time.sleep(1)

    with yaspin(text="Сохранение") as spinner:
        merged_wb.save(save_path)
        spinner.text = 'Сохранено'
        spinner.ok('✔️')
        time.sleep(1)
    print("🔄  Форматирование строк")

    total_rows = output_ws.max_row - 1
    for row in tqdm(output_ws.iter_rows(min_row=2, max_row=output_ws.max_row), total=total_rows, desc="🔄  Форматирование строк", bar_format=bar, ncols=150, leave=False, ascii=' ▱▰'):
        for col_idx, cell in enumerate(row, 1):
            val_str = str(cell.value) if cell.value is not None else ''
            cell.alignment = Alignment(horizontal='left')
            column_widths[col_idx] = max(column_widths.get(col_idx, 0), len(val_str) + 10)

        # Добавление ссылок
        if len(row) >= 11:
            title_cell = row[1]  # 2-я колонка - Название
            url_cell = row[10]  # 11-я колонка - URL
            offer_url = str(url_cell.value)

            if offer_url:
                title_cell.value = f'=HYPERLINK("{offer_url}", "{title_cell.value}")'
                title_cell.font = hlink_style

                url_cell.hyperlink = offer_url
                url_cell.font = hlink_style

    # Установка ширины колонок
    for col_idx, width in column_widths.items():
        output_ws.column_dimensions[get_column_letter(col_idx)].width = min(width, 100)

    print("\r✅  Форматирование завершено", end="", flush=True)
    time.sleep(1)

    try:
        with yaspin(text="Сохранение") as spinner:
            merged_wb.save(save_path)
            spinner.text = 'Сохранено'
            spinner.ok('✔️')
            time.sleep(1)

        print(f"💾  {LIGHT_GREEN}Файл сохранен в `{clickable_file_link(save_path)}` {WHITE}")
        time.sleep(1)

    except PermissionError as e:
        logger.error(f"Не удалось сохранить EXCEL файл: {e}")


def register_styles(wb: Workbook):
    """
    Регистрирует пользовательские стили в рабочей книге Excel, если они ещё не добавлены.

    Проверяет наличие каждого стиля в списке зарегистрированных именованных стилей рабочей книги `wb`.
    Если стиль отсутствует, добавляет его в рабочую книгу.

    :param wb: Объект рабочей книги Excel (`openpyxl.Workbook`), в которую нужно добавить стили.
    """

    def add_style(style):
        if style.name not in wb.named_styles:
            wb.add_named_style(style)

    add_style(success_status)
    add_style(error_status)
    add_style(hidden_status)
    add_style(not_specified_status)


def save_offers(content: list[Offer], region_id, region_name, city_id, city_name, category_id, category_name, out_dir, save_json, save_xls):
    """
    Сохраняет список предложений (`Offer`) в указанные форматы (JSON и/или Excel) по заданной иерархии директорий.

    Формирует имя файла на основе ID и названия категории, создает директорию с учетом региона и города,
    а затем сохраняет данные в формате JSON и/или XLSX, если соответствующие флаги активированы.

    :param content: Список предложений для сохранения.
    :param region_id: Идентификатор региона.
    :param region_name: Название региона.
    :param city_id: Идентификатор города.
    :param city_name: Название города.
    :param category_id: Идентификатор категории.
    :param category_name: Название категории.
    :param out_dir: Базовая директория для сохранения файлов.
    :param save_json: Флаг, указывающий, нужно ли сохранять данные в формате JSON.
    :param save_xls: Флаг, указывающий, нужно ли сохранять данные в формате XLSX.
    """
    filename = validate_filename(f'{region_id}_{city_id}_{category_id}_{category_name}__offers({len(content)})')
    file_path = os.path.join(out_dir, f"{region_name.replace(' ', '-')}_{region_id}", f"{city_name}_{city_id}")
    os.makedirs(file_path, exist_ok=True)

    if save_json:
        save_json([item.model_dump() for item in content], os.path.join(out_dir, f'{filename}.json'))

    if save_xls:
        save_offers_excel(content, os.path.join(file_path, f'{filename}.xlsx'), show_info=False)


async def process_cell(parser, n, item, total, counter, ws, wb, wb_path, save_every_n=20):
    """
    Обрабатывает ячейку таблицы.
    Получает ячейку с данными об объявлении, потом получает номер телефона с OLX, и записывает телефон обратно в ячейку

    :param parser: olxParser
    :param n: Текущая итерация
    :param item: Данные объявления
    :param total: Общее количество обхявлений в файле
    :param counter: Счетчик обраьотки ячеек
    :param ws: Рабочий лист
    :param wb: Рабочая книга
    :param wb_path: Путь до файла
    :param save_every_n: Сохранение файла каждые x итераций
    :return:
    """
    async with lock:
        counter['value'] += 1
        progress = f"[{counter['value']} / {total}]"

    offer_id = item[0]
    url = item[10]
    row_idx = n + 2
    number_cell = ws.cell(row=row_idx, column=3)
    digits = ''.join(re.findall(r'\d+', str(item[2])))

    # Если номер не указан, удален, скрыт или уже получен, то пропускаем ячейку
    if number_cell.value == 'False':
        number_cell.value = 'не указан'
        number_cell.style = 'not_specified_status'
        print(f"{progress}  ℹ  {DARK_GRAY}  Номер не указан {''.ljust(20)}{WHITE} · {url}")
        return

    if number_cell.value == 'не указан':
        return

    if number_cell.value == 'удален':
        return

    if number_cell.value == 'скрыт':
        return

    if digits.isdigit():
        return

    response = await parser.get_phone_number(offer_id, response_only=True)
    if isinstance(response, Exception):
        number_cell.value = 'ошибка'
        number_cell.style = 'error_status'
        print(f"{progress}  ❌{RED}  Номер не получен: {WHITE}Ошибка при получении номера: {response} · {url}")
        return
    if response == {}:
        number_cell.value = 'Captcha'
        number_cell.style = 'error_status'
        print(f"{progress}  ❌{RED}  Номер не получен: {WHITE}Captcha {''.ljust(9)} ·  {url}")
        return

    # Если ответ получен и есть ошибка, то ставим соответствующий статус в ячейку
    if 'error' in response:
        error = response.get('error', {}).get('detail')
        if error == 'Disallowed for this user':
            number_cell.value = 'скрыт'
            number_cell.style = 'hidden_status'
        elif error == 'Ad is not active':
            number_cell.value = 'удален'
            number_cell.style = 'not_specified_status'
        elif 'Невозможно продолжить' in error:
            number_cell.value = 'Captcha'
            number_cell.style = 'error_status'
            print(f"{progress}  ❌{RED}  Номер не получен: {WHITE} Captcha {''.ljust(9)} · {url}")
        else:
            error = error.split('.')[0]
            print(f"{progress}  ❌{RED}  Номер не получен: {WHITE} {error} · {url}")

        # Если номер был скрыт, то пытаемся его получить, если все успешно то добавляем его в ячейку
        if number_cell.value == 'скрыт':
            phone = await parser.get_phone_number(offer_id)
            if phone:
                number_cell.value = phone
                number_cell.style = 'success_status'
                print(f"{progress}  ✔{GREEN}  Номер получен: {LIGHT_YELLOW}{phone.ljust(20)}{WHITE} · {url}")
            else:
                print(f"{progress}  ❌{RED}  Номер не получен: {WHITE} Ошибка 1 · {url}")

    else:
        # Если ошибок нет, то получаем номер из ответа, создаем строку из номеров и записываем в ячейку
        phones = response.get('data', {}).get('phones')

        if phones:
            phone = ' · '.join([str(p) for p in phones])
            number_cell.value = phone
            number_cell.style = 'success_status'
            print(f"{progress}  ✔{LIGHT_GREEN}  Номер получен: {LIGHT_YELLOW}{phone.ljust(20)}{WHITE} · {url}")
        else:
            print(f"{progress}  ❌{DARK_GRAY}  Номер не указан {''.ljust(20)}{WHITE} · {url}")

    # Если во время получения номера произошла ошибка или капча, то пытаемся еще раз получить номер
    if number_cell.value == 'True' or number_cell.value == 'Captcha':
        phone = await parser.get_phone_number(offer_id)
        if phone:
            number_cell.value = phone
            number_cell.style = 'success_status'
            print(f"{progress}  ✔{LIGHT_CYAN}  Номер получен: {LIGHT_YELLOW}{phone.ljust(20)}{WHITE} · {url}")
        else:
            print(f"{progress}  ❌{RED}  Номер не получен: {WHITE} Ошибка 2 · {url}")

    # Сохраняем прогресс каждые N итераций
    if (n + 1) % save_every_n == 0:
        async with lock:
            wb.save(wb_path)
        get_token(exp_time_only=True)
